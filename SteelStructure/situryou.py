import openpyxl
from openpyxl import load_workbook

# 既存のワークブックを読み込む
wb = load_workbook('点検歩廊.xlsx')
ws = wb.active

# セルの値を取得
cell_value = ws['A1'].value
print(f"A1 セルの値: {cell_value}")